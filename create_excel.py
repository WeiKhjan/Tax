import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os

# ============================================
# PBC LIST EXCEL FILE
# ============================================

wb_pbc = openpyxl.Workbook()
ws_pbc = wb_pbc.active
ws_pbc.title = "PBC List"

# Define styles
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
section_font = Font(bold=True, color="FFFFFF", size=11)
section_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
high_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
medium_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
low_fill = PatternFill(start_color="69DB7C", end_color="69DB7C", fill_type="solid")
received_fill = PatternFill(start_color="D3F9D8", end_color="D3F9D8", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
title_font = Font(bold=True, size=16, color="1F4E79")
subtitle_font = Font(bold=True, size=12)

# Title section
ws_pbc.merge_cells('A1:E1')
ws_pbc['A1'] = "PBC LIST (PROVIDED BY CLIENT)"
ws_pbc['A1'].font = title_font
ws_pbc['A1'].alignment = Alignment(horizontal='center')

ws_pbc.merge_cells('A2:E2')
ws_pbc['A2'] = "JATI KIRANA SDN BHD"
ws_pbc['A2'].font = subtitle_font
ws_pbc['A2'].alignment = Alignment(horizontal='center')

ws_pbc.merge_cells('A3:E3')
ws_pbc['A3'] = "Year of Assessment 2025 | Basis Period: 01 July 2024 to 30 June 2025"
ws_pbc['A3'].alignment = Alignment(horizontal='center')

# Headers
headers = ['No.', 'Document Description', 'Priority', 'Status', 'Remarks']
for col, header in enumerate(headers, 1):
    cell = ws_pbc.cell(row=5, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

# PBC Data
pbc_data = [
    ("", "A. STATUTORY & CORPORATE DOCUMENTS", "", "", ""),
    ("A1", "SSM Company Profile / Business Registration", "HIGH", "Pending", "Required for company details"),
    ("A2", "Certificate of Incorporation", "Medium", "Pending", ""),
    ("A3", "Memorandum & Articles of Association (M&A)", "Medium", "Pending", ""),
    ("A4", "Latest Annual Return (AR)", "HIGH", "Pending", "Required for shareholder details"),
    ("A5", "Register of Directors", "HIGH", "Pending", "Required for Form C"),
    ("A6", "Register of Shareholders", "HIGH", "Pending", "Required for SME verification"),
    ("A7", "Register of Beneficial Owners", "HIGH", "Pending", "Required for Form C"),
    ("A8", "Board Resolutions (relevant to tax matters)", "Low", "Pending", "If applicable"),
    ("", "B. FINANCIAL STATEMENTS & ACCOUNTS", "", "", ""),
    ("B1", "Audited Financial Statements (current year)", "N/A", "Received", "Management accounts received"),
    ("B2", "Audited Financial Statements (prior year)", "Medium", "Pending", "For comparison"),
    ("B3", "Trial Balance (detailed)", "N/A", "Received", ""),
    ("B4", "General Ledger", "N/A", "Received", ""),
    ("B5", "Bank Statements (12 months)", "Low", "Pending", "For verification"),
    ("", "C. FIXED ASSETS & CAPITAL ALLOWANCES", "", "", ""),
    ("C1", "Fixed Asset Register (detailed with acquisition dates)", "HIGH", "Pending", "Required for CA computation"),
    ("C2", "Invoices for asset additions (current year)", "N/A", "N/A", "No additions"),
    ("C3", "Asset disposal documents", "N/A", "N/A", "No disposals"),
    ("C4", "Hire purchase agreements", "Low", "Pending", "If applicable"),
    ("C5", "Prior year capital allowance schedule", "HIGH", "Pending", "Required for TWDV B/F"),
    ("C6", "Prior year tax computation", "HIGH", "Pending", "Required for losses B/F"),
    ("", "D. INCOME RELATED DOCUMENTS", "", "", ""),
    ("D1", "Rental agreement - A12 Senibong Cove", "Medium", "Pending", "Tenancy with Unigra Food"),
    ("D2", "Property Sale & Purchase Agreement", "Medium", "Pending", "For cost verification"),
    ("", "E. TAX COMPLIANCE DOCUMENTS", "", "", ""),
    ("E1", "Prior year Form C and tax computation", "HIGH", "Pending", "Critical for tax position"),
    ("E2", "CP204 submission for current year", "Medium", "Pending", "For verification"),
    ("E3", "LHDN correspondence / Notices", "Low", "Pending", "If any"),
    ("", "F. RELATED PARTY DOCUMENTS", "", "", ""),
    ("F1", "List of related companies", "Medium", "Pending", "[Related Company] relationship"),
    ("F2", "Directors loan agreements (if any)", "Medium", "Pending", "Interest-free advances"),
    ("F3", "Intercompany transaction listing", "Medium", "Pending", ""),
]

row = 6
for item in pbc_data:
    if item[1].startswith(("A.", "B.", "C.", "D.", "E.", "F.")):
        # Section header
        ws_pbc.merge_cells(f'A{row}:E{row}')
        cell = ws_pbc.cell(row=row, column=1, value=item[1])
        cell.font = section_font
        cell.fill = section_fill
        cell.border = thin_border
    else:
        for col, value in enumerate(item, 1):
            cell = ws_pbc.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

            # Color coding for priority
            if col == 3:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if value == "HIGH":
                    cell.fill = high_fill
                    cell.font = Font(bold=True)
                elif value == "Medium":
                    cell.fill = medium_fill
                elif value == "Low":
                    cell.fill = low_fill

            # Color coding for status
            if col == 4:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if value == "Received":
                    cell.fill = received_fill
    row += 1

# Legend
row += 1
ws_pbc.cell(row=row, column=1, value="PRIORITY LEGEND:").font = Font(bold=True)
row += 1
ws_pbc.cell(row=row, column=1, value="HIGH").fill = high_fill
ws_pbc.cell(row=row, column=2, value="Critical for tax computation and Form C filing")
row += 1
ws_pbc.cell(row=row, column=1, value="Medium").fill = medium_fill
ws_pbc.cell(row=row, column=2, value="Important for verification and disclosure")
row += 1
ws_pbc.cell(row=row, column=1, value="Low").fill = low_fill
ws_pbc.cell(row=row, column=2, value="Supporting documentation")

# Column widths
ws_pbc.column_dimensions['A'].width = 8
ws_pbc.column_dimensions['B'].width = 50
ws_pbc.column_dimensions['C'].width = 12
ws_pbc.column_dimensions['D'].width = 12
ws_pbc.column_dimensions['E'].width = 35

# Save PBC file
pbc_path = r"C:\Users\khjan\Downloads\Demo - YYC - Calude\JATI KIRANA SDN BHD YA 2025\PBC_Outstanding_Items.xlsx"
wb_pbc.save(pbc_path)
print(f"Created: {pbc_path}")

# ============================================
# QUERY LIST EXCEL FILE
# ============================================

wb_qry = openpyxl.Workbook()
ws_qry = wb_qry.active
ws_qry.title = "Query List"

# Title section
ws_qry.merge_cells('A1:F1')
ws_qry['A1'] = "TAX QUERY LIST"
ws_qry['A1'].font = title_font
ws_qry['A1'].alignment = Alignment(horizontal='center')

ws_qry.merge_cells('A2:F2')
ws_qry['A2'] = "JATI KIRANA SDN BHD"
ws_qry['A2'].font = subtitle_font
ws_qry['A2'].alignment = Alignment(horizontal='center')

ws_qry.merge_cells('A3:F3')
ws_qry['A3'] = "Year of Assessment 2025 | Basis Period: 01 July 2024 to 30 June 2025"
ws_qry['A3'].alignment = Alignment(horizontal='center')

# Headers
qry_headers = ['Query No.', 'Reference / Account', 'Amount (RM)', 'Query Description', 'Tax Implication', 'Client Response']
for col, header in enumerate(qry_headers, 1):
    cell = ws_qry.cell(row=5, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Query Data
query_data = [
    ("Q001", "A/C 5100/0000\nRental Income", "43,600.00",
     "Please confirm if there was rental income for May and June 2025.\n\nBased on the general ledger, rental receipts were only recorded from July 2024 to April 2025 (10 months only).\n\nWas the property vacant for 2 months, or is there outstanding rental receivable?",
     "Rental income must be recognized on an accrual basis under Section 4(d) ITA 1967.\n\nIf rental was receivable but not received, it should still be included as taxable income.",
     ""),
    ("Q002", "A/C 3501/0000\n[Related Company] Sdn Bhd", "170,000.00",
     "Please clarify:\n\n1. Is this an interest-free loan?\n2. What is the relationship between Jati Kirana and [Related Company]?\n3. Is [Related Company] a shareholder of Jati Kirana?\n4. What is [Related Company]'s paid-up capital?",
     "IMPORTANT: If [Related Company]'s paid-up capital exceeds RM 2.5 million, Jati Kirana will NOT qualify for SME tax rates.\n\nIf related party, transactions must be at arm's length (Section 140A).",
     ""),
    ("Q003", "A/C 4100/0001 & 4100/0002\nAmounts Owing to Directors", "69,923.87",
     "Please confirm:\n\n1. Are the directors' advances interest-free?\n2. Is there any formal loan agreement?\n3. When are these amounts intended to be repaid?\n\nDirectors:\n- [Director 1]: RM 39,923.87\n- [Director 2]: RM 30,000.00",
     "Under Section 140B ITA 1967, loans from directors may be subject to deemed interest adjustment if the company has interest-bearing borrowings while receiving interest-free loans.",
     ""),
    ("Q004", "A/C 4200/1003\n[Creditor Name]", "30,300.00",
     "Please clarify:\n\n1. Who/what is [Creditor Name]?\n2. What is the nature of this liability?\n3. Is this a related party?",
     "Related party transactions must be disclosed in Form C.\n\nNeed to determine relationship for transfer pricing compliance.",
     ""),
    ("Q005", "A/C 4200/1001\n[Creditor Name]", "70,676.99",
     "Please confirm:\n\n1. Who is [Creditor Name] and what is the relationship to the company/directors?\n2. What is the nature of this amount owing?\n3. Is [Creditor Name] a shareholder of the company?",
     "Related party transactions must be disclosed.\n\nIf shareholder, may affect beneficial ownership disclosure requirements.",
     ""),
    ("Q006", "Fixed Assets\nInvestment Property - A12 Senibong Cove", "583,800.00",
     "Please provide:\n\n1. Date of acquisition of the property\n2. Copy of Sale & Purchase Agreement\n3. Breakdown of acquisition cost:\n   - Purchase price\n   - Stamp duty\n   - Legal fees\n   - Renovation (if any)",
     "Required to verify cost basis and determine if any portion qualifies for capital allowance (e.g., fixtures included in purchase).",
     ""),
]

row = 6
for item in query_data:
    for col, value in enumerate(item, 1):
        cell = ws_qry.cell(row=row, column=col, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='top', wrap_text=True)

        if col == 1:  # Query No
            cell.alignment = Alignment(horizontal='center', vertical='top')
            cell.font = Font(bold=True)
        elif col == 3:  # Amount
            cell.alignment = Alignment(horizontal='right', vertical='top')
        elif col == 6:  # Client Response - highlight for input
            cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

    ws_qry.row_dimensions[row].height = 120
    row += 1

# Instructions section
row += 2
ws_qry.merge_cells(f'A{row}:F{row}')
ws_qry.cell(row=row, column=1, value="INSTRUCTIONS:").font = Font(bold=True, size=11)
row += 1
instructions = [
    "1. Please provide written response to each query in the 'Client Response' column (highlighted in yellow)",
    "2. Attach supporting documents where applicable",
    "3. Sign and return this form upon completion"
]
for inst in instructions:
    ws_qry.merge_cells(f'A{row}:F{row}')
    ws_qry.cell(row=row, column=1, value=inst)
    row += 1

# Signature section
row += 2
ws_qry.cell(row=row, column=1, value="Responded By:").font = Font(bold=True)
ws_qry.cell(row=row, column=2, value="_________________________")
ws_qry.cell(row=row, column=4, value="Date:").font = Font(bold=True)
ws_qry.cell(row=row, column=5, value="_________________________")
row += 2
ws_qry.cell(row=row, column=1, value="Signature:").font = Font(bold=True)
ws_qry.cell(row=row, column=2, value="_________________________")

# Column widths
ws_qry.column_dimensions['A'].width = 12
ws_qry.column_dimensions['B'].width = 22
ws_qry.column_dimensions['C'].width = 15
ws_qry.column_dimensions['D'].width = 45
ws_qry.column_dimensions['E'].width = 40
ws_qry.column_dimensions['F'].width = 30

# Save Query file
qry_path = r"C:\Users\khjan\Downloads\Demo - YYC - Calude\JATI KIRANA SDN BHD YA 2025\Query_List.xlsx"
wb_qry.save(qry_path)
print(f"Created: {qry_path}")

# Remove old CSV files
csv_pbc = r"C:\Users\khjan\Downloads\Demo - YYC - Calude\JATI KIRANA SDN BHD YA 2025\PBC_Outstanding_Items.csv"
csv_qry = r"C:\Users\khjan\Downloads\Demo - YYC - Calude\JATI KIRANA SDN BHD YA 2025\Query_List.csv"
if os.path.exists(csv_pbc):
    os.remove(csv_pbc)
    print(f"Removed: {csv_pbc}")
if os.path.exists(csv_qry):
    os.remove(csv_qry)
    print(f"Removed: {csv_qry}")

print("\nDone! Excel files created with proper formatting.")
