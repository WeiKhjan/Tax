"""
Tax Computation 2021 - Kilat Sejadi Sdn Bhd
Creates Tax Computation 2021 Excel file matching YA2020 structure.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

import sys
OUTPUT_PATH = sys.argv[1] if len(sys.argv) > 1 else "Tax_Computation.xlsx"

# ── Style helpers ─────────────────────────────────────────────────────────────

def _border(**kwargs):
    sides = {k: Side(style=v) for k, v in kwargs.items() if v}
    return Border(**sides)

def _font(bold=False, italic=False, size=10, underline=None):
    return Font(bold=bold, italic=italic, size=size, underline=underline)

def _align(h=None, v=None, wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def w(ws, row, col, value,
      bold=False, italic=False, size=10,
      halign=None, wrap=False,
      bot=None, top=None, left=None, right=None,
      nfmt=None):
    """Write a cell with optional styling."""
    cell = ws.cell(row=row, column=col, value=value)
    if bold or italic or size != 10:
        cell.font = _font(bold=bold, italic=italic, size=size)
    if halign or wrap:
        cell.alignment = _align(h=halign, wrap=wrap)
    borders = {}
    if bot:   borders['bottom'] = Side(style=bot)
    if top:   borders['top']    = Side(style=top)
    if left:  borders['left']   = Side(style=left)
    if right: borders['right']  = Side(style=right)
    if borders:
        cell.border = Border(**borders)
    if nfmt:
        cell.number_format = nfmt
    return cell

def num(ws, row, col, value,
        bold=False, bot=None, top=None, nfmt='#,##0.00'):
    """Write a numeric cell (right-aligned)."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = _align(h='right')
    if bold:
        cell.font = _font(bold=True)
    borders = {}
    if bot: borders['bottom'] = Side(style=bot)
    if top: borders['top']    = Side(style=top)
    if borders:
        cell.border = Border(**borders)
    cell.number_format = nfmt
    return cell

def formula(ws, row, col, expr,
            bold=False, bot=None, top=None, nfmt='#,##0.00'):
    """Write a formula cell (right-aligned)."""
    return num(ws, row, col, expr, bold=bold, bot=bot, top=top, nfmt=nfmt)

def col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def row_height(ws, row, height):
    ws.row_dimensions[row].height = height

# ── Sheet 1: YA2021 (Cover / Index) ──────────────────────────────────────────

def build_ya2021(wb):
    ws = wb.create_sheet("YA2021")

    # Column widths
    col_width(ws, 'A', 12)
    col_width(ws, 'B', 4)
    col_width(ws, 'C', 3)
    col_width(ws, 'D', 40)
    col_width(ws, 'E', 10)
    col_width(ws, 'F', 10)
    col_width(ws, 'G', 10)
    col_width(ws, 'H', 10)

    # Header
    w(ws, 1, 1, 'KILAT SEJADI SDN. BHD.', bold=True, size=12)
    w(ws, 2, 1, 'YEAR OF ASSESSMENT 2021', halign='left')
    w(ws, 3, 1, 'BASIS PERIOD : FOR THE FINANCIAL YEAR ENDED 31 DECEMBER 2021', halign='left')
    w(ws, 4, 1, 'TAX REF : [TAX REF NO]', halign='left')

    # Index title
    w(ws, 9, 4, 'I  N  D  E  X', halign='center', bold=True, size=12)

    # Statements
    w(ws, 14, 1, 'Statement', halign='left')
    w(ws, 14, 2, 'A', halign='center')
    w(ws, 14, 3, ':')
    w(ws, 14, 4, 'Income Tax Computation')
    w(ws, 15, 2, 'B', halign='center')
    w(ws, 15, 3, ':')
    w(ws, 15, 4, 'Capital Allowance Schedule')

    # Schedules
    w(ws, 17, 1, 'Schedule')
    w(ws, 17, 2, 'A', halign='center')
    w(ws, 17, 3, ':')
    w(ws, 17, 4, 'Fine, penalty and late charges')
    w(ws, 18, 4, "Auditors' remuneration")
    w(ws, 19, 4, 'Professional fees (accounting fee)')
    w(ws, 20, 4, 'Tax agent fee')

# ── Sheet 2: St.A (Main Tax Computation) ─────────────────────────────────────

def build_st_a(wb):
    ws = wb.create_sheet("St.A")

    # Column widths (match YA2020 approx)
    col_width(ws, 'A', 6)
    col_width(ws, 'B', 6)
    col_width(ws, 'C', 38)
    col_width(ws, 'D', 4)
    col_width(ws, 'E', 4)
    col_width(ws, 'F', 4)
    col_width(ws, 'G', 14)
    col_width(ws, 'H', 4)
    col_width(ws, 'I', 14)
    col_width(ws, 'J', 4)
    col_width(ws, 'K', 10)

    # Row heights for key rows
    for r in [1, 2, 3, 4, 5]:
        row_height(ws, r, 15)

    # ── Header ─────────────────────────────────────────────────────────────
    w(ws, 1, 1, 'KILAT SEJADI SDN. BHD.', bold=True)
    w(ws, 1, 9, 'Statement A', halign='right')
    w(ws, 2, 1, 'INCOME TAX COMPUTATION')
    w(ws, 3, 1, 'BASIS PERIOD: 1 JAN 2021 TO 31 DEC 2021', halign='left')
    w(ws, 4, 1, 'YEAR OF ASSESSMENT 2021', halign='left')
    w(ws, 5, 1, 'TAX REF : [TAX REF NO]', halign='left')

    # ── Column Headers ─────────────────────────────────────────────────────
    w(ws, 9, 1, '64923 Licensed money lending activities')
    w(ws, 9, 7, 'RM', halign='center')
    w(ws, 9, 9, 'RM', halign='center')
    w(ws, 9, 11, 'REF', halign='left')

    # ── Profit / (Loss) before tax ─────────────────────────────────────────
    w(ws, 11, 1, 'Profit / (Loss) before tax', halign='left')
    num(ws, 11, 9, -17347)
    w(ws, 11, 11, 'P & L')

    # ── Add: Non-deductible expenses ───────────────────────────────────────
    w(ws, 13, 1, 'Add  ', halign='left')
    w(ws, 13, 2, ':',    halign='left')
    w(ws, 13, 3, 'Non-deductible expenses')

    # Row 14 - Audit fee
    w(ws, 14, 3, "Auditors' remuneration")
    num(ws, 14, 7, 3000)
    w(ws, 14, 11, 'P & L')

    # Row 15 - Depreciation
    w(ws, 15, 3, 'Depreciation of plant and equipment', halign='left')
    num(ws, 15, 7, 3836)
    w(ws, 15, 11, 'P & L')

    # Row 16 - Fines
    w(ws, 16, 3, 'Fine, penalty and late charges')
    formula(ws, 16, 7, "='Sch.A'!G16")
    w(ws, 16, 11, 'SCH. A')

    # Row 17 - Professional fee
    w(ws, 17, 3, 'Professional fee (accounting fee)')
    formula(ws, 17, 7, "='Sch.A'!G26")
    w(ws, 17, 11, 'SCH. A')

    # Row 18 - Tax agent fee
    w(ws, 18, 3, 'Tax agent fee')
    formula(ws, 18, 7, "='Sch.A'!G31")
    w(ws, 18, 11, 'SCH. A')

    # Row 19 - Subtotal add-backs (single underline top + bottom)
    formula(ws, 19, 9, '=SUM(G14:G18)', bot='thin', top='thin')

    # Row 21 - Income before further deductions
    formula(ws, 21, 9, '=SUM(I11,I19)')

    # ── Less: Further deduction ────────────────────────────────────────────
    w(ws, 23, 1, 'Less',  halign='left')
    w(ws, 23, 2, ':',     halign='left')
    w(ws, 23, 3, 'Further deduction')

    # Row 24 - Audit re-allowed
    w(ws, 24, 3, "Auditors' remuneration  [S.33(1)(a)]")
    formula(ws, 24, 7, '=-G14')
    w(ws, 24, 11, 'SCH. A')

    # Row 25 - Professional fee re-allowed
    w(ws, 25, 3, 'Professional fee - accounting fee  [S.33(1)]')
    formula(ws, 25, 7, '=-G17')
    w(ws, 25, 11, 'SCH. A')

    # Row 26 - Tax agent re-allowed
    w(ws, 26, 3, 'Tax agent fee  [S.34B]')
    formula(ws, 26, 7, '=-G18')
    w(ws, 26, 11, 'SCH. A')

    # Row 27 - Subtotal further deductions (single underline)
    formula(ws, 27, 9, '=SUM(G24:G26)', bot='thin', top='thin')

    # ── Adjusted income / (loss) ───────────────────────────────────────────
    w(ws, 29, 1, 'Adjusted income / (loss)', halign='left')
    formula(ws, 29, 9, '=SUM(I21,I27)', bot='thin')

    # ── Capital Allowance ──────────────────────────────────────────────────
    w(ws, 31, 1, 'Less',  halign='left')
    w(ws, 31, 2, ':',     halign='left')
    w(ws, 31, 3, 'Capital allowance')
    num(ws, 31, 9, 0)
    w(ws, 31, 11, 'St. B')

    # ── Statutory income / (loss) ──────────────────────────────────────────
    w(ws, 33, 1, 'Statutory income / (loss)', halign='left')
    formula(ws, 33, 9, '=SUM(I29:I32)', bot='thin')

    # ── Tax loss b/f ───────────────────────────────────────────────────────
    w(ws, 35, 1, 'Less',  halign='left')
    w(ws, 35, 2, ':',     halign='left')
    w(ws, 35, 3, 'Tax loss b/f')
    num(ws, 35, 9, 0)

    # ── Aggregate income / (loss) ──────────────────────────────────────────
    w(ws, 37, 1, 'Aggregate income / (loss)', halign='left')
    formula(ws, 37, 9, '=SUM(I33:I36)', bot='thin')

    # ── Current year business loss ─────────────────────────────────────────
    w(ws, 39, 1, 'Less',  halign='left')
    w(ws, 39, 2, ':',     halign='left')
    w(ws, 39, 3, 'Current year business loss')
    num(ws, 39, 9, 0)

    # ── Chargeable income ──────────────────────────────────────────────────
    w(ws, 41, 1, 'Chargeable income', halign='left', bold=True)
    formula(ws, 41, 9, '=SUM(I37:I40)', bold=True, bot='double', top='thin')

    # ── Tax Payable ────────────────────────────────────────────────────────
    w(ws, 43, 1, 'Income tax payable', halign='left')
    w(ws, 44, 1, 'Tax payable at 17%', halign='left')
    formula(ws, 44, 9, '=IF(I41>0,I41*0.17,0)', bot='thin')

    # ── CP204 ──────────────────────────────────────────────────────────────
    w(ws, 46, 1, 'Less: CP204 paid', halign='left')
    num(ws, 46, 9, 0)

    # ── Total tax due ──────────────────────────────────────────────────────
    w(ws, 48, 1, 'Total tax due / (refundable)', halign='left', bold=True)
    formula(ws, 48, 9, '=I44+I46', bold=True, bot='double', top='thin')

    # ── Loss carried forward note ──────────────────────────────────────────
    w(ws, 51, 1, 'Note: Tax loss carried forward to YA2022', italic=True)
    w(ws, 52, 3, 'Adjusted loss YA2021')
    num(ws, 52, 9, -13461, bot='thin')
    w(ws, 52, 11, 'St. A')
    w(ws, 53, 3, 'Add: Tax loss b/f')
    num(ws, 53, 9, 0)
    w(ws, 54, 3, 'Tax loss c/f to YA2022', bold=True)
    formula(ws, 54, 9, '=ABS(I52)+I53', bold=True, bot='double', top='thin')

# ── Sheet 3: Sch.A (Supporting Schedule) ─────────────────────────────────────

def build_sch_a(wb):
    ws = wb.create_sheet("Sch.A")

    # Column widths
    col_width(ws, 'A', 52)
    col_width(ws, 'B', 4)
    col_width(ws, 'C', 4)
    col_width(ws, 'D', 4)
    col_width(ws, 'E', 14)
    col_width(ws, 'F', 4)
    col_width(ws, 'G', 14)
    col_width(ws, 'H', 8)

    # ── Header ─────────────────────────────────────────────────────────────
    w(ws, 1, 1, 'KILAT SEJADI SDN. BHD.', bold=True)
    w(ws, 1, 7, 'Schedule A', halign='right')
    w(ws, 3, 1, 'BASIS PERIOD : 1 JAN 2021 TO 31 DEC 2021', halign='left')
    w(ws, 4, 1, 'YEAR OF ASSESSMENT 2021', halign='left')
    w(ws, 5, 1, 'TAX REF : [TAX REF NO]', halign='left')

    # Column headers
    w(ws, 6, 7, 'DISALLOWED', halign='center')
    w(ws, 7, 5, 'RM', halign='center')
    w(ws, 7, 7, 'RM', halign='center')

    # ── Section 1: Fine, Penalty and Late Charges ──────────────────────────
    w(ws, 9, 1, 'FINE, PENALTY AND LATE CHARGES', bold=True)

    # Items
    FINES = [
        (11, '08 FEB 2021 - SOCSO PENALTY',  10),
        (12, '08 FEB 2021 - EIS PENALTY',     10),
        (13, '15 APR 2021 - EIS PENALTY',     10),
        (14, '15 APR 2021 - SOCSO PENALTY',   10),
        (15, '12 MAY 2021 - SOCSO PENALTY',   10),
    ]
    for row, desc, amt in FINES:
        w(ws, row, 1, desc)
        num(ws, row, 5, amt, nfmt='#,##0.00')
        formula(ws, row, 7, f'=E{row}')

    # Total fines
    formula(ws, 16, 5, '=SUM(E11:E15)', bot='thin', top='thin')
    formula(ws, 16, 7, '=SUM(G11:G15)', bot='thin', top='thin')
    w(ws, 16, 8, 'St. A', halign='center')

    # ── Section 2: Auditors' Remuneration ─────────────────────────────────
    w(ws, 18, 1, "AUDITORS' REMUNERATION", bold=True)
    w(ws, 19, 5, 'RM', halign='center')

    w(ws, 20, 1, "AUDIT FEE FY2021 (ACCRUED)")
    num(ws, 20, 5, 3000, nfmt='#,##0.00')
    formula(ws, 20, 7, '=E20')

    formula(ws, 21, 5, '=SUM(E20)', bot='thin', top='thin')
    formula(ws, 21, 7, '=SUM(G20)', bot='thin', top='thin')
    w(ws, 21, 8, 'St. A', halign='center')

    # ── Section 3: Professional Fees - Accounting Fee ─────────────────────
    w(ws, 23, 1, 'PROFESSIONAL FEES - ACCOUNTING FEE', bold=True)
    w(ws, 24, 5, 'RM', halign='center')

    w(ws, 25, 1, 'ACCOUNTING FEE FY2021 (ACCRUED)')
    num(ws, 25, 5, 2000, nfmt='#,##0.00')
    formula(ws, 25, 7, '=E25')

    formula(ws, 26, 5, '=SUM(E25)', bot='thin', top='thin')
    formula(ws, 26, 7, '=SUM(G25)', bot='thin', top='thin')
    w(ws, 26, 8, 'St. A', halign='center')

    # ── Section 4: Tax Agent Fee ───────────────────────────────────────────
    w(ws, 28, 1, 'TAX AGENT FEE', bold=True)
    w(ws, 29, 5, 'RM', halign='center')

    w(ws, 30, 1, 'TAX AGENT FEE YA2021 (ACCRUED)')
    num(ws, 30, 5, 1500, nfmt='#,##0.00')
    formula(ws, 30, 7, '=E30')

    formula(ws, 31, 5, '=SUM(E30)', bot='thin', top='thin')
    formula(ws, 31, 7, '=SUM(G30)', bot='thin', top='thin')
    w(ws, 31, 8, 'St. A', halign='center')

# ── Sheet 4: St.B (Capital Allowance) ────────────────────────────────────────

def build_st_b(wb):
    ws = wb.create_sheet("St.B")

    # Column widths
    col_width(ws, 'A', 4)
    col_width(ws, 'B', 42)
    col_width(ws, 'C', 6)
    col_width(ws, 'D', 6)
    col_width(ws, 'E', 6)
    col_width(ws, 'F', 4)
    col_width(ws, 'G', 11)
    col_width(ws, 'H', 11)
    col_width(ws, 'I', 11)
    col_width(ws, 'J', 11)
    col_width(ws, 'K', 11)
    col_width(ws, 'L', 11)
    col_width(ws, 'M', 11)

    # ── Header ─────────────────────────────────────────────────────────────
    w(ws, 1, 1, 'KILAT SEJADI SDN. BHD.', bold=True)
    w(ws, 1, 13, 'Statement B', halign='right')
    w(ws, 2, 1, 'CAPITAL ALLOWANCE SCHEDULE')
    w(ws, 3, 1, 'YEAR OF ASSESSMENT 2021', halign='left')
    w(ws, 4, 1, 'TAX REF : [TAX REF NO]', halign='left')

    # Column headers
    w(ws, 6, 2, 'Capital Allowance')
    w(ws, 7, 7, 'YA2021', halign='center')
    w(ws, 8, 3, 'YEAR', halign='center')
    w(ws, 8, 4, 'IA%', halign='center')
    w(ws, 8, 5, 'AA%', halign='center')
    w(ws, 8, 7, 'COST', halign='center')
    w(ws, 8, 8, 'QE', halign='center')
    w(ws, 8, 9, 'TWDV B/F', halign='center')
    w(ws, 8, 10, 'ADDITION', halign='center')
    w(ws, 8, 11, 'IA', halign='center')
    w(ws, 8, 12, 'AA', halign='center')
    w(ws, 8, 13, 'TWDV C/F', halign='center')

    # Apply thin borders to header row
    for col in range(3, 14):
        ws.cell(row=8, column=col).border = Border(
            bottom=Side(style='thin'), top=Side(style='thin'))

    # ── Asset categories with TWDV B/F = 0, No additions ──────────────────
    # All assets were fully written off by YA2020. TWDV B/F = 0 for all.
    # No additions in 2021. CA = NIL.

    # Helper: add asset category
    def add_section(header_row, header, assets, total_row):
        """assets: list of (desc, year, ia_pct, aa_pct, cost, qe)"""
        w(ws, header_row, 2, header)
        for i, (desc, year, ia, aa, cost, qe) in enumerate(assets):
            r = header_row + 1 + i
            w(ws, r, 2, desc)
            if year:  num(ws, r, 3, year, nfmt='0')
            if ia is not None: num(ws, r, 4, ia, nfmt='0%')
            if aa is not None: num(ws, r, 5, aa, nfmt='0%')
            num(ws, r, 7, cost)
            num(ws, r, 8, qe)
            num(ws, r, 9, 0)   # TWDV B/F = 0
            num(ws, r, 10, 0)  # Addition = 0
            num(ws, r, 11, 0)  # IA = 0
            num(ws, r, 12, 0)  # AA = 0
            num(ws, r, 13, 0)  # TWDV C/F = 0

        # Subtotal row with SUM formulas
        start = header_row + 1
        end   = total_row - 1
        for col in [7, 8, 9, 10, 11, 12, 13]:
            cl = get_column_letter(col)
            formula(ws, total_row, col, f'=SUM({cl}{start}:{cl}{end})',
                    bot='thin', top='thin')

    # ── 1. Computer Equipment ──────────────────────────────────────────────
    add_section(10, 'COMPUTER EQUIPMENT', [
        ('1 unit of 19" LCD Monitor',                    2016, 0.2, 0.1, 150, 150),
        ('1 unit of Asus A507M-ABR061T Notebook',        2019, 0.2, 0.8, 1299, 1299),
        ('1 unit of Asus K31AN-Bing-MY002S',             2015, 0.2, 0.1, 999, 999),
    ], total_row=14)

    # ── 2. Electrical Installation ─────────────────────────────────────────
    add_section(16, 'ELECTRICAL INSTALLATION', [
        ('Electrical installation',                       2015, 0, 0, 2020, 0),
    ], total_row=18)

    # ── 3. Furniture and Fitting ───────────────────────────────────────────
    add_section(20, 'FURNITURE AND FITTING', [
        ('1 unit of Mirror',                             2015, 0.2, 0.1, 280, 280),
        ('1 unit of Sofa',                               2015, 0.2, 0.1, 1690, 1690),
        ('1 unit of Table',                              2015, 0.2, 0.1, 2396, 2396),
        ('1 unit of Table, chair and sofa',              2015, 0.2, 0.1, 1648, 1648),
    ], total_row=25)

    # ── 4. Office Equipment ────────────────────────────────────────────────
    add_section(27, 'OFFICE EQUIPMENT', [
        ('1 set of Alarm - Door access V2000, 8zone Alarm set, CCTV wiring, Y3 Dome camera',
                                                          2015, 0.2, 0.1, 2735, 2735),
        ('1 unit of LS 5 safety box',                    2015, 0.2, 0.1, 3590, 3590),
        ("1 unit of Oshino 18' Floor fan",               2015, 0.2, 0.1, 170, 170),
        ('1 unit of Panasonic 32 TV, wet & dry v/cleaner',
                                                          2015, 0.2, 0.1, 1184, 1184),
        ('1 unit of Printer HP-Laser Jet',               2015, 0.2, 0.1, 1479, 1479),
        ('1 unit of Printer HP2135 Deskjet printer Aio', 2015, 0.2, 0.1, 199, 199),
        ('1 unit of Samsung Galaxy Core Prime SM-G361HZWVXME (W)',
                                                          2015, 0.2, 0.1, 529, 529),
        ('1 unit of York Aircon',                        2015, 0.2, 0.1, 2800, 2800),
        ('1 unit of York Aircon',                        2015, 0.2, 0.1, 2200, 2200),
    ], total_row=37)

    # ── 5. Renovation ─────────────────────────────────────────────────────
    add_section(39, 'RENOVATION', [
        ('Renovation',                                   2015, 0, 0, 10450, 0),
    ], total_row=41)

    # ── 6. Revenue Item Capitalised ────────────────────────────────────────
    add_section(43, 'REVENUE ITEM CAPITALISED', [
        ('1 set of Handphone Oppo A83',                  2018, 0.2, 0.8, 680, 680),
        ('Small value asset',                            2017, 0.2, 0.8, 1225, 1225),
    ], total_row=46)

    # ── 7. Signboard ──────────────────────────────────────────────────────
    add_section(48, 'SIGNBOARD', [
        ('1 unit of Signboard - Light Box, glass sticker', 2015, 0, 0, 2544, 0),
    ], total_row=50)

    # ── Grand Total ────────────────────────────────────────────────────────
    for col in [7, 8, 9, 10, 11, 12, 13]:
        cl = get_column_letter(col)
        formula(ws, 52, col, f'={cl}14+{cl}18+{cl}25+{cl}37+{cl}41+{cl}46+{cl}50',
                bold=True, bot='double', top='thin')

    w(ws, 52, 2, 'TOTAL', bold=True)

    # ── CA brought forward note ────────────────────────────────────────────
    w(ws, 54, 2, 'Note: All assets fully depreciated / written off. TWDV b/f = NIL.',
      italic=True)
    w(ws, 55, 2, 'No additions in YA2021. Capital Allowance claim = RM NIL.',
      italic=True)

# ── Sheet 5: DF (Deferred Tax Computation) ────────────────────────────────────

def build_df(wb):
    ws = wb.create_sheet("DF")

    # Column widths
    col_width(ws, 'A', 4)
    col_width(ws, 'B', 42)
    col_width(ws, 'C', 4)
    col_width(ws, 'D', 14)
    col_width(ws, 'E', 14)
    col_width(ws, 'F', 14)
    col_width(ws, 'G', 14)
    col_width(ws, 'H', 14)

    # ── Header ─────────────────────────────────────────────────────────────
    w(ws, 1, 1, 'KILAT SEJADI SDN. BHD.', bold=True)
    w(ws, 3, 1, 'YEAR OF ASSESSMENT 2021', halign='left')
    w(ws, 4, 1, 'SUBJECT: DEFERRED TAX COMPUTATION', halign='left')
    w(ws, 5, 1, 'TAX REF : [TAX REF NO]', halign='left')

    w(ws, 7, 2, 'Deferred tax')
    w(ws, 8, 6, 'Temporary Difference', bold=True, halign='center')

    # Column headers for deferred tax
    w(ws, 9, 2, '2021', halign='left')
    w(ws, 9, 4, 'NBV', bold=True)
    w(ws, 9, 5, 'TWDV C/F', bold=True)
    w(ws, 9, 6, 'Temp. Diff.', bold=True, halign='center')

    # ── Qualifying assets ──────────────────────────────────────────────────
    w(ws, 10, 2, 'Qualifying assets (Plant & equipment)')
    num(ws, 10, 4, 12077)    # NBV per audit report
    num(ws, 10, 5, 0)        # TWDV (tax) = 0 (fully claimed)
    formula(ws, 10, 6, '=D10-E10')  # = 12,077

    # No hire purchase in 2021

    # ── DTL on plant & equipment ───────────────────────────────────────────
    w(ws, 12, 2, 'Taxable temporary difference (plant & equipment)')
    formula(ws, 12, 6, '=F10')  # 12,077

    w(ws, 13, 2, 'Tax rate')
    num(ws, 13, 5, 0.17, nfmt='0%')
    w(ws, 14, 2, 'Deferred tax liability (DTL) on plant & equipment')
    formula(ws, 14, 6, '=F12*E13')  # 12,077 × 17% = 2,053

    # ── DTA on tax loss ────────────────────────────────────────────────────
    w(ws, 16, 2, 'Less: Tax loss carried forward (YA2021)')
    num(ws, 16, 4, 13461)    # Unabsorbed business loss
    formula(ws, 16, 6, '=-D16*E13')   # (2,286)

    # ── Net position ───────────────────────────────────────────────────────
    w(ws, 18, 2, 'Net deferred tax asset / (liability)', bold=True)
    formula(ws, 18, 6, '=F14+F16', bold=True, bot='thin', top='thin')
    # Result: 2,053 - 2,286 = (233)

    # ── Going concern / recognition note ──────────────────────────────────
    w(ws, 20, 2,
      "Note: The net deferred tax asset of RM233 has NOT been recognised as",
      italic=True)
    w(ws, 21, 2,
      "the company is a going concern with uncertainties about future profits.",
      italic=True)
    w(ws, 22, 2,
      "DTL is fully reversed in YA2021 as plant & equipment TWDV = NIL.",
      italic=True)

    w(ws, 24, 2, 'Recognised DTL / (DTA) at 31 December 2021', bold=True)
    num(ws, 24, 6, 0, bold=True, bot='double', top='thin')

    # ── Summary ────────────────────────────────────────────────────────────
    w(ws, 26, 2, 'Movement in deferred tax liability:')
    w(ws, 27, 2, 'Opening balance (DTL per audit - YA2020)')
    num(ws, 27, 6, 0)
    w(ws, 28, 2, 'Reversal / (Charge) during YA2021')
    num(ws, 28, 6, 0)
    w(ws, 29, 2, 'Closing balance at 31 Dec 2021', bold=True)
    formula(ws, 29, 6, '=F27+F28', bold=True, bot='double', top='thin')


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_ya2021(wb)
    build_st_a(wb)
    build_sch_a(wb)
    build_st_b(wb)
    build_df(wb)

    wb.save(OUTPUT_PATH)
    print(f"[OK] Saved: {OUTPUT_PATH}")

    # Quick sanity checks
    wb2 = openpyxl.load_workbook(OUTPUT_PATH, data_only=False)
    print(f"   Sheets: {wb2.sheetnames}")
    for sname in wb2.sheetnames:
        ws = wb2[sname]
        cells_with_data = sum(1 for row in ws.iter_rows() for c in row if c.value is not None)
        print(f"   {sname}: {cells_with_data} cells populated")

if __name__ == '__main__':
    main()
