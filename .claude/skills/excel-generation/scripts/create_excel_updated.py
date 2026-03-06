import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Styles
header_font = Font(bold=True, size=14, color="FFFFFF")
normal_font = Font(size=10)
header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
section_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
received_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
pending_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
resolved_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Create Updated PBC Checklist Excel
wb_pbc = openpyxl.Workbook()
ws_pbc = wb_pbc.active
ws_pbc.title = "PBC Checklist"

# Header
ws_pbc.merge_cells('A1:F1')
ws_pbc['A1'] = "PBC LIST - CIRCULAR TECH ASIA SDN. BHD. (202401008787)"
ws_pbc['A1'].font = header_font
ws_pbc['A1'].fill = header_fill
ws_pbc['A1'].alignment = center_align
ws_pbc.row_dimensions[1].height = 30

ws_pbc.merge_cells('A2:F2')
ws_pbc['A2'] = "YA 2025 | Basis Period: 01/03/2024 to 28/02/2025 | STATUS: SUBSTANTIALLY COMPLETE"
ws_pbc['A2'].font = Font(size=10, italic=True, bold=True)
ws_pbc['A2'].alignment = center_align

# Column headers
headers = ['No', 'Document Description', 'Status', 'Date Received', 'Details/Remarks', 'Source File']
col_widths = [6, 45, 12, 14, 35, 25]
for col, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws_pbc.cell(row=4, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws_pbc.column_dimensions[get_column_letter(col)].width = width

# PBC Items - Updated with received status
pbc_items = [
    ("A. STATUTORY & CORPORATE DOCUMENTS", [
        ("A1", "SSM Company Profile / Business Registration", "RECEIVED", "Feb 2025", "Co. No: 202401008787 (1554637-X)", "Section 14.pdf"),
        ("A2", "Certificate of Incorporation", "RECEIVED", "Feb 2025", "Incorporated: 05/03/2024", "Section 14.pdf"),
        ("A3", "Register of Directors", "RECEIVED", "Feb 2025", "FOO MUN YEE - 720512-08-5788", "Section 14.pdf"),
        ("A4", "Register of Shareholders / Members", "RECEIVED", "Feb 2025", "FOO MUN YEE - 100 shares (100%)", "Section 14.pdf"),
        ("A5", "Register of Beneficial Owners", "RECEIVED", "Feb 2025", "FOO MUN YEE - 100% Direct", "CTASB-BO.pdf"),
    ]),
    ("B. TAX REGISTRATION", [
        ("B1", "LHDN Tax Registration", "RECEIVED", "Feb 2025", "Tax File: C 58577123050", "LHDN EMAIL.pdf"),
        ("B2", "Employer Registration", "RECEIVED", "Feb 2025", "Employer No: E 9608867307", "LHDN EMAIL.pdf"),
        ("B3", "CP204 Submission for YA 2025", "PENDING", "", "To verify if submitted", ""),
    ]),
    ("C. DIRECTORS INFORMATION (Schedule C)", [
        ("C1", "Director's full name", "RECEIVED", "Feb 2025", "FOO MUN YEE", "Section 14.pdf"),
        ("C2", "Director's NRIC (12 digits)", "RECEIVED", "Feb 2025", "720512-08-5788", "Section 14.pdf"),
        ("C3", "Director's residential address", "RECEIVED", "Feb 2025", "46 Jalan SS 21/25, Damansara Utama, 47400 PJ", "CTASB-Director.pdf"),
        ("C4", "Director's date of appointment", "RECEIVED", "Feb 2025", "05 March 2024", "CTASB-Director.pdf"),
        ("C5", "Director's email", "RECEIVED", "Feb 2025", "lydiafoo@hotmail.com", "Section 14.pdf"),
    ]),
    ("D. SHAREHOLDERS INFORMATION (Schedule D)", [
        ("D1", "List of all shareholders", "RECEIVED", "Feb 2025", "FOO MUN YEE - 100%", "Section 14.pdf"),
        ("D2", "Shareholder NRIC", "RECEIVED", "Feb 2025", "720512-08-5788", "Section 14.pdf"),
        ("D3", "Corporate shareholder details", "RECEIVED", "Feb 2025", "NONE - 100% individual", "Section 14.pdf"),
    ]),
    ("E. BENEFICIAL OWNERSHIP (Schedule E)", [
        ("E1", "BO Declaration Form", "RECEIVED", "Feb 2025", "Section 56 declaration", "CTASB-BO.pdf"),
        ("E2", "BO Name & NRIC", "RECEIVED", "Feb 2025", "FOO MUN YEE - 720512-08-5788", "CTASB-BO.pdf"),
        ("E3", "Date became BO", "RECEIVED", "Feb 2025", "05 March 2024", "CTASB-BO.pdf"),
    ]),
]

row = 5
for section_name, items in pbc_items:
    ws_pbc.merge_cells(f'A{row}:F{row}')
    ws_pbc[f'A{row}'] = section_name
    ws_pbc[f'A{row}'].font = Font(bold=True, color="FFFFFF")
    ws_pbc[f'A{row}'].fill = section_fill
    ws_pbc[f'A{row}'].alignment = left_align
    row += 1

    for item in items:
        no, desc, status, date, details, source = item
        ws_pbc.cell(row=row, column=1, value=no).alignment = center_align
        ws_pbc.cell(row=row, column=2, value=desc).alignment = left_align
        ws_pbc.cell(row=row, column=3, value=status).alignment = center_align
        ws_pbc.cell(row=row, column=4, value=date).alignment = center_align
        ws_pbc.cell(row=row, column=5, value=details).alignment = left_align
        ws_pbc.cell(row=row, column=6, value=source).alignment = left_align

        for col in range(1, 7):
            ws_pbc.cell(row=row, column=col).border = thin_border
            ws_pbc.cell(row=row, column=col).font = normal_font

        if status == "RECEIVED":
            for col in range(1, 7):
                ws_pbc.cell(row=row, column=col).fill = received_fill
        elif status == "PENDING":
            for col in range(1, 7):
                ws_pbc.cell(row=row, column=col).fill = pending_fill
        row += 1

# Summary
row += 1
ws_pbc[f'A{row}'] = "SUMMARY: All critical documents received. Tax working papers updated. Query Q001 resolved."
ws_pbc[f'A{row}'].font = Font(bold=True)
row += 1
ws_pbc[f'A{row}'] = "Outstanding: Q002 (Phone bill), Q003 (Prepaid expenses), Q004 (T-shirt) - Minor items, no material impact"

wb_pbc.save("C:/Users/khjan/Downloads/Demo - YYC - Calude/Circular Tech Asia Sdn Bhd YA 2025/PBC_Checklist_CircularTech_YA2025_UPDATED.xlsx")
print("Updated PBC Checklist Excel created successfully")

# Create Updated Query List Excel
wb_qry = openpyxl.Workbook()
ws_qry = wb_qry.active
ws_qry.title = "Query List"

# Header
ws_qry.merge_cells('A1:F1')
ws_qry['A1'] = "TAX QUERY LIST - CIRCULAR TECH ASIA SDN. BHD. (202401008787)"
ws_qry['A1'].font = header_font
ws_qry['A1'].fill = header_fill
ws_qry['A1'].alignment = center_align
ws_qry.row_dimensions[1].height = 30

ws_qry.merge_cells('A2:F2')
ws_qry['A2'] = "YA 2025 | Q001 RESOLVED | Q002-Q004 Pending (Minor - No Material Impact)"
ws_qry['A2'].font = Font(size=10, italic=True, bold=True)
ws_qry['A2'].alignment = center_align

# Column headers
qry_headers = ['Query No', 'Reference', 'Amount (RM)', 'Query Description', 'Status', 'Resolution/Response']
qry_widths = [10, 18, 12, 40, 12, 45]
for col, (header, width) in enumerate(zip(qry_headers, qry_widths), 1):
    cell = ws_qry.cell(row=4, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws_qry.column_dimensions[get_column_letter(col)].width = width

# Query items
queries = [
    ("Q001", "Director Fee\n(Acc 6524)", "10,000.00",
     "Director particulars required:\n- Full name\n- NRIC\n- Address\n- Date of appointment",
     "RESOLVED",
     "FOO MUN YEE\nNRIC: 720512-08-5788\nAddress: 46 Jalan SS 21/25, Damansara Utama, 47400 PJ\nAppointed: 05/03/2024\nFee FULLY DEDUCTIBLE"),
    ("Q002", "Telephone\n(Acc 6513)", "1,349.11",
     "Phone bills to Celcom Axiata (12 months)\n\n1. Is phone 100% business use?\n2. Registered under company/personal?\n3. Any private use element?",
     "PENDING",
     "[Awaiting client response]\n\nIf private element: Max add-back RM 674.56 (50%)"),
    ("Q003", "Prepaid Exp\n(Acc 1005)", "1,965.60",
     "Prepaid Professional Fee 2025\n\n1. What period does this cover?\n2. Nature of fee?",
     "PENDING",
     "[For information only]\n\nNo tax impact - already excluded from deduction"),
    ("Q004", "General Exp\n(Acc 6508)", "228.00",
     "T-Shirt Printing (PrintCious)\n\n1. Purpose of t-shirts?\n2. Promotional items?\n3. Staff uniforms?",
     "PENDING",
     "[Awaiting client response]\n\nIf personal: Add-back RM 228"),
]

row = 5
for qno, ref, amt, desc, status, resolution in queries:
    ws_qry.cell(row=row, column=1, value=qno).alignment = center_align
    ws_qry.cell(row=row, column=2, value=ref).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws_qry.cell(row=row, column=3, value=amt).alignment = Alignment(horizontal='right', vertical='top')
    ws_qry.cell(row=row, column=4, value=desc).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws_qry.cell(row=row, column=5, value=status).alignment = center_align
    ws_qry.cell(row=row, column=6, value=resolution).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    for col in range(1, 7):
        ws_qry.cell(row=row, column=col).border = thin_border
        ws_qry.cell(row=row, column=col).font = normal_font

    if status == "RESOLVED":
        for col in range(1, 7):
            ws_qry.cell(row=row, column=col).fill = resolved_fill
    elif status == "PENDING":
        for col in range(1, 7):
            ws_qry.cell(row=row, column=col).fill = pending_fill

    ws_qry.row_dimensions[row].height = 90
    row += 1

# Summary
row += 1
ws_qry[f'A{row}'] = "IMPACT ASSESSMENT:"
ws_qry[f'A{row}'].font = Font(bold=True)
row += 1
ws_qry[f'A{row}'] = "Even if all pending queries (Q002-Q004) result in add-backs, maximum impact is RM 902.56."
row += 1
ws_qry[f'A{row}'] = "Company remains in LOSS position with NIL TAX PAYABLE regardless of query outcomes."
row += 1
ws_qry[f'A{row}'] = "Recommendation: Tax computation can be finalized. Follow up Q002-Q004 for documentation completeness."

wb_qry.save("C:/Users/khjan/Downloads/Demo - YYC - Calude/Circular Tech Asia Sdn Bhd YA 2025/Query_List_CircularTech_YA2025_UPDATED.xlsx")
print("Updated Query List Excel created successfully")
