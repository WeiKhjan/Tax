import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# ============================================
# UPDATE PBC LIST - MARK CA SCHEDULE RECEIVED
# ============================================

import sys
output_dir = sys.argv[1] if len(sys.argv) > 1 else "."
pbc_path = f"{output_dir}/PBC_Outstanding_Items.xlsx"
wb_pbc = openpyxl.load_workbook(pbc_path)
ws_pbc = wb_pbc.active

# Find and update C5 (Prior year capital allowance schedule) row
# Based on the data structure, it should be around row 22-23
received_fill = PatternFill(start_color="D3F9D8", end_color="D3F9D8", fill_type="solid")

# Iterate through rows to find "Prior year capital allowance schedule"
for row in range(6, 40):
    cell_a = ws_pbc.cell(row=row, column=1)
    cell_b = ws_pbc.cell(row=row, column=2)
    if cell_a.value and "C5" in str(cell_a.value):
        # Update status to Received
        ws_pbc.cell(row=row, column=4, value="Received")
        ws_pbc.cell(row=row, column=4).fill = received_fill
        ws_pbc.cell(row=row, column=5, value="TWDV = NIL, Total Assets RM 607,776")
        print(f"Updated C5 at row {row}")
        break
    if cell_b.value and "capital allowance schedule" in str(cell_b.value).lower():
        ws_pbc.cell(row=row, column=4, value="Received")
        ws_pbc.cell(row=row, column=4).fill = received_fill
        ws_pbc.cell(row=row, column=5, value="TWDV = NIL, Total Assets RM 607,776")
        print(f"Updated CA schedule at row {row}")
        break

wb_pbc.save(pbc_path)
print(f"Updated: {pbc_path}")

# ============================================
# UPDATE QUERY LIST - ADD CONFIRMED INFO TO Q006
# ============================================

qry_path = f"{output_dir}/Query_List.xlsx"
wb_qry = openpyxl.load_workbook(qry_path)
ws_qry = wb_qry.active

# Define styles
title_font = Font(bold=True, size=16, color="1F4E79")
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Update Q006 query description with confirmed info from CA schedule
for row in range(6, 15):
    cell_a = ws_qry.cell(row=row, column=1)
    if cell_a.value == "Q006":
        # Update the query description to show what we now know
        ws_qry.cell(row=row, column=4, value="""Per CA Schedule received:
- Property: Lot No. A1-22 Block A, No.1 Persiaran Senibong 81750 Masai, Johor
- Property Name: The WaterEdge Residence, Senibong Cove
- Purchase Year: 2016
- Purchase Price (per CA Schedule): RM 576,794.00
- Per GL (A/C 2009/0000): RM 583,800.00
- Difference: RM 7,006.00

QUERY: Please clarify the RM 7,006 difference between CA schedule and GL.
Is this due to renovation, stamp duty, or legal fees?

Also please provide:
- Copy of Sale & Purchase Agreement
- Breakdown of acquisition cost""")
        print(f"Updated Q006 at row {row}")
        break

wb_qry.save(qry_path)
print(f"Updated: {qry_path}")

print("\nDone! Excel files updated with CA schedule information.")
